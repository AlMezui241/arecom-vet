from django.shortcuts import render, redirect
from django.views import View
from django.views.generic import ListView, DetailView, CreateView, UpdateView, TemplateView
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.urls import reverse_lazy
from django.db.models import Sum, Q
from django.http import HttpResponse
from django.utils import timezone
from decimal import Decimal
import csv

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

from .models import Activite, TypeActivite, ActiviteAuditLog
from .forms import ActiviteForm, ActiviteDocumentFormSet

from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm
from reportlab.lib import colors
from io import BytesIO


class ExportActivitePDFView(LoginRequiredMixin, View):
    """Génération d'un reçu/ordre de recette au format PDF pour une Activité (Identique à VET)"""

    def get(self, request, numero, *args, **kwargs):
        try:
            activite = Activite.objects.select_related('type_activite').get(numero=numero)

            buffer = BytesIO()
            p = canvas.Canvas(buffer, pagesize=A4)
            width, height = A4

            # Configuration des polices (Standard)
            font_bold = "Helvetica-Bold"
            font_reg = "Helvetica"

            # En-tête (République du Gabon)
            p.setFont(font_bold, 16)
            p.drawCentredString(width / 2, height - 2 * cm, "RÉPUBLIQUE DU GABON")
            p.setFont(font_reg, 10)
            p.drawCentredString(width / 2, height - 2.5 * cm, "Union - Travail - Justice")
            p.line(width / 2 - 2 * cm, height - 2.7 * cm, width / 2 + 2 * cm, height - 2.7 * cm)

            # Titre du document
            p.setFont(font_bold, 14)
            p.drawCentredString(width / 2, height - 4.5 * cm, f"ORDRE DE RECETTE - {activite.type_activite.nom.upper()}")
            p.setFont(font_reg, 11)
            p.drawCentredString(width / 2, height - 5 * cm, f"N° {activite.numero}")

            # Infos Établissement
            p.setFont(font_bold, 12)
            p.drawString(2 * cm, height - 7 * cm, "INFORMATIONS DE L'ÉTABLISSEMENT")
            p.setFont(font_reg, 11)
            p.drawString(2 * cm, height - 7.7 * cm, f"Raison Sociale : {activite.nom_activite}")
            p.drawString(2 * cm, height - 8.4 * cm, f"Localisation : {activite.quartier_zone}, {activite.localite}, {activite.province}")
            p.drawString(2 * cm, height - 9.1 * cm, f"Téléphone : {activite.telephone or 'N/A'}")

            # Détails Financiers (Tableau Exhaustif comme VET)
            p.setFont(font_bold, 12)
            p.drawString(2 * cm, height - 10.5 * cm, "DÉTAILS DES FRAIS")

            y = height - 11.2 * cm
            p.setFont(font_bold, 10)
            p.drawString(2 * cm, y, "N°")
            p.drawString(3 * cm, y, "Désignation")
            p.drawRightString(width - 2 * cm, y, "Montant (FCFA)")
            p.line(2 * cm, y - 0.2 * cm, width - 2 * cm, y - 0.2 * cm)

            p.setFont(font_reg, 10)
            idx = 1
            if activite.type_activite.a_frais_instruction:
                y -= 0.8 * cm
                p.drawString(2 * cm, y, str(idx))
                p.drawString(3 * cm, y, "Frais d'instruction de dossier")
                p.drawRightString(width - 2 * cm, y, f"{float(activite.frais_instruction_dossier or 0):,.0f}")
                idx += 1

            if activite.type_activite.a_frais_exploitation:
                y -= 0.6 * cm
                p.drawString(2 * cm, y, str(idx))
                p.drawString(3 * cm, y, "Frais annuels d'exploitation")
                p.drawRightString(width - 2 * cm, y, f"{float(activite.frais_exploitation_annuel or 0):,.0f}")
                idx += 1

            if activite.type_activite.a_contribution_fonds:
                y -= 0.6 * cm
                p.drawString(2 * cm, y, str(idx))
                p.drawString(3 * cm, y, "Contribution annuelle au fonds universel")
                p.drawRightString(width - 2 * cm, y, f"{float(activite.contribution_fonds_universel or 0):,.0f}")
                idx += 1

            p.line(width - 5 * cm, y - 0.4 * cm, width - 2 * cm, y - 0.4 * cm)
            y -= 1 * cm
            p.setFont(font_bold, 12)
            p.drawString(2 * cm, y, "TOTAL À RECOUVRER")
            p.drawRightString(width - 2 * cm, y, f"{float(activite.montant_total_redevance):,.0f} FCFA")

            # Statuts de conformité et paiement
            y -= 1.5 * cm
            p.setFont(font_bold, 11)
            p.drawString(2 * cm, y, "CONFORMITÉ ET PAIEMENT :")
            p.setFont(font_reg, 10)
            y -= 0.6 * cm
            p.drawString(2.5 * cm, y, f"Facture d'achat présente : {'OUI' if activite.presence_facture else 'NON'}")
            y -= 0.5 * cm
            p.drawString(2.5 * cm, y, f"Autorisation présente : {'OUI' if activite.presence_autorisation else 'NON'}")
            y -= 0.5 * cm
            p.drawString(2.5 * cm, y, f"Frais de dossier : {'RÉGLÉS' if activite.frais_de_dossier_payes else 'NON RÉGLÉS'}")
            y -= 0.5 * cm
            p.drawString(2.5 * cm, y, f"Redevance Annuelle : {'RÉGLÉE' if activite.redevance_payee else 'NON RÉGLÉE'}")

            # Pied de page
            p.setFont("Helvetica-Oblique", 9)
            p.drawCentredString(width / 2, 2 * cm,
                                f"Document généré le {timezone.now().strftime('%d/%m/%Y à %H:%M')} - ARCEP")

            p.showPage()
            p.save()

            buffer.seek(0)
            filename = f"recu_{activite.numero}.pdf"
            return FileResponse(buffer, as_attachment=True, filename=filename, content_type='application/pdf')
        except Activite.DoesNotExist:
            return HttpResponse("Entité non trouvée", status=404)
        except Exception as e:
            return HttpResponse(f"Erreur lors de la génération du PDF : {str(e)}", status=500)


class ActiviteFilterMixin:
    """Partage la logique de filtrage pour les querysets d'Activite."""
    def get_filtered_queryset(self):
        qs = Activite.objects.select_related('type_activite').all()
        q = self.request.GET.get("q")
        if q:
            qs = qs.filter(
                Q(nom_activite__icontains=q)
                | Q(numero__icontains=q)
                | Q(province__icontains=q)
                | Q(localite__icontains=q)
                | Q(quartier_zone__icontains=q)
            )
        type_slug = self.request.GET.get('type')
        if type_slug:
            qs = qs.filter(type_activite__slug=type_slug)
        province = self.request.GET.get('province')
        if province:
            qs = qs.filter(province=province)
        statut = self.request.GET.get('statut')
        if statut:
            qs = qs.filter(statut=statut)
        frais = self.request.GET.get('frais')
        if frais == "paye":
            qs = qs.filter(frais_de_dossier_payes=True)
        elif frais == "non_paye":
            qs = qs.filter(frais_de_dossier_payes=False)
        redevance = self.request.GET.get('redevance')
        if redevance == "paye":
            qs = qs.filter(redevance_payee=True)
        elif redevance == "non_paye":
            qs = qs.filter(redevance_payee=False)
        return qs


class ActiviteListView(LoginRequiredMixin, ActiviteFilterMixin, ListView):
    """Liste toutes les entités avec recherche et filtres (inspiré de VETListView)."""
    model = Activite
    template_name = 'activites/activites_list.html'
    context_object_name = 'activites'
    paginate_by = 25

    def get_queryset(self):
        return self.get_filtered_queryset()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['types'] = TypeActivite.objects.filter(actif=True)
        context['provinces'] = (
            Activite.objects.order_by('province')
            .values_list('province', flat=True)
            .distinct()
        )
        return context


class ActiviteDetailView(LoginRequiredMixin, DetailView):
    """Détail d'une entité"""
    model = Activite
    template_name = 'activites/activites_detail.html'
    context_object_name = 'activite'
    slug_field = 'numero'
    slug_url_kwarg = 'numero'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        activite = self.get_object()
        context['documents'] = activite.documents.all()
        context['today'] = timezone.now().date()
        return context


class ActiviteCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    permission_required = 'activites.add_activite'
    model = Activite
    form_class = ActiviteForm
    template_name = 'activites/activites_form.html'
    success_url = reverse_lazy('activites:list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        if self.request.POST:
            context['documents'] = ActiviteDocumentFormSet(
                self.request.POST,
                self.request.FILES
            )
        else:
            context['documents'] = ActiviteDocumentFormSet()
            
        return context

    def form_valid(self, form):
        context = self.get_context_data()
        documents = context['documents']
        if documents.is_valid():
            self.object = form.save(commit=False)
            self.object.created_by = self.request.user
            self.object.save()
            documents.instance = self.object
            documents.save()
            ActiviteAuditLog.objects.create(
                activite=self.object,
                user=self.request.user,
                action='create',
                details=f"Création de l'activité {self.object.numero}"
            )
            return super().form_valid(form)
        else:
            return self.render_to_response(self.get_context_data(form=form))




class ActiviteUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    permission_required = 'activites.change_activite'
    model = Activite
    form_class = ActiviteForm
    template_name = 'activites/activites_form.html'
    slug_field = 'numero'
    slug_url_kwarg = 'numero'
    success_url = reverse_lazy('activites:list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        if self.request.POST:
            context['documents'] = ActiviteDocumentFormSet(
                self.request.POST,
                self.request.FILES,
                instance=self.object
            )
        else:
            context['documents'] = ActiviteDocumentFormSet(instance=self.object)
            
        return context

    def form_valid(self, form):
        context = self.get_context_data()
        documents = context['documents']
        if documents.is_valid():
            self.object = form.save()
            documents.instance = self.object
            documents.save()
            ActiviteAuditLog.objects.create(
                activite=self.object,
                user=self.request.user,
                action='update',
                details=f"Modification de l'activité {self.object.numero}"
            )
            return super().form_valid(form)
        else:
            return self.render_to_response(self.get_context_data(form=form))




def activites_dashboard(request):
    """Dashboard global des entités"""
    if not request.user.is_authenticated:
        return redirect('admin:login')

    stats = {
        'total_activites': Activite.objects.count(),
        'par_type': {},
        'par_province': {},
        'total_redevances': Decimal('0.00'),
    }

    # Stats par type
    for type_activite in TypeActivite.objects.filter(actif=True):
        activites_type = Activite.objects.filter(type_activite=type_activite)
        stats['par_type'][type_activite.nom] = {
            'count': activites_type.count(),
            'redevances': activites_type.aggregate(
                Sum('frais_exploitation_annuel'))['frais_exploitation_annuel__sum'] or Decimal('0.00')
        }

    # Stats par province
    for province in Activite.objects.values_list('province', flat=True).distinct():
        activites_province = Activite.objects.filter(province=province)
        stats['par_province'][province] = activites_province.count()

    total_redevances = sum(a.montant_total_redevance for a in Activite.objects.all())
    stats['total_redevances'] = total_redevances

    return render(request, 'activites/dashboard.html', {'stats': stats})


class ExportActiviteExcelListView(LoginRequiredMixin, ActiviteFilterMixin, View):
    """Export des entités au format Excel (.xlsx) en respectant les filtres de la liste."""

    def get(self, request, *args, **kwargs):
        qs = self.get_filtered_queryset()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Activites"

        headers = [
            "Numéro",
            "Nom / Raison sociale",
            "Type",
            "Province",
            "Localité",
            "Quartier/Zone",
            "Frais instruction",
            "Frais exploitation",
            "Contribution fonds",
            "Total redevances",
            "Frais dossier payés",
            "Redevance payée",
            "Statut",
        ]
        ws.append(headers)

        header_fill = PatternFill(start_color="0D6EFD", end_color="0D6EFD", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for e in qs:
            ws.append(
                [
                    e.numero,
                    e.nom_activite,
                    e.type_activite.nom,
                    e.province,
                    e.localite,
                    e.quartier_zone,
                    e.frais_instruction_dossier or Decimal("0"),
                    e.frais_exploitation_annuel or Decimal("0"),
                    e.contribution_fonds_universel or Decimal("0"),
                    e.montant_total_redevance or Decimal("0"),
                    "Oui" if e.frais_de_dossier_payes else "Non",
                    "Oui" if e.redevance_payee else "Non",
                    e.get_statut_display(),
                ]
            )

        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            ws.column_dimensions[column].width = max_length + 2

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        filename = f"activites_{timezone.now().strftime('%Y%m%d')}.xlsx"
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response


class ExportActiviteCSVView(LoginRequiredMixin, PermissionRequiredMixin, ActiviteFilterMixin, View):
    permission_required = 'activites.view_activite'
    """Export des entités au format CSV."""

    def get(self, request, *args, **kwargs):
        qs = self.get_filtered_queryset()

        response = HttpResponse(content_type="text/csv; charset=utf-8")
        filename = f"activites_{timezone.now().strftime('%Y%m%d_%H%M')}.csv"
        response["Content-Disposition"] = f'attachment; filename="{filename}"'

        writer = csv.writer(response, delimiter=";")
        writer.writerow(
            [
                "Numéro",
                "Nom / Raison sociale",
                "Type",
                "Province",
                "Localité",
                "Quartier/Zone",
                "Frais instruction",
                "Frais exploitation",
                "Contribution fonds",
                "Total redevances",
                "Frais dossier payés",
                "Redevance payée",
                "Statut",
            ]
        )

        for e in qs:
            writer.writerow(
                [
                    e.numero,
                    e.nom_activite,
                    e.type_activite.nom,
                    e.province,
                    e.localite,
                    e.quartier_zone,
                    e.frais_instruction_dossier or Decimal("0"),
                    e.frais_exploitation_annuel or Decimal("0"),
                    e.contribution_fonds_universel or Decimal("0"),
                    e.montant_total_redevance or Decimal("0"),
                    "OUI" if e.frais_de_dossier_payes else "NON",
                    "OUI" if e.redevance_payee else "NON",
                    e.get_statut_display(),
                ]
            )

        return response


class ActiviteMapView(LoginRequiredMixin, TemplateView):
    """Carte interactive des entités avec coordonnées GPS."""

    template_name = "activites/activites_map.html"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)

        activites_qs = (
            Activite.objects.exclude(latitude__isnull=True)
            .exclude(longitude__isnull=True)
            .select_related("type_activite")
        )

        activite_points = []
        for e in activites_qs:
            activite_points.append(
                {
                    "numero": e.numero,
                    "nom_activite": e.nom_activite,
                    "latitude": float(e.latitude),
                    "longitude": float(e.longitude),
                    "province": e.province,
                    "localite": e.localite,
                    "quartier_zone": e.quartier_zone,
                    "statut": e.statut,
                    "frais_de_dossier_payes": e.frais_de_dossier_payes,
                    "redevance_payee": e.redevance_payee,
                    "montant_total_redevance": float(e.montant_total_redevance or 0),
                    "dossier_suivi_par": e.dossier_suivi_par,
                    "type_activite": e.type_activite.nom,
                }
            )

        ctx["activite_points"] = activite_points
        ctx["provinces"] = (
            Activite.objects.order_by("province").values_list("province", flat=True).distinct()
        )
        return ctx

