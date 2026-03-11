from django.db.models import Q, Sum, Count, F, Value, DecimalField
from django.urls import reverse_lazy
from django.utils import timezone
from datetime import timedelta
from django.views import View
from django.http import HttpResponse, HttpResponseForbidden, FileResponse
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView, TemplateView
from django.forms import inlineformset_factory
from .models import VET, FRAIS_DOSSIER, VETVignette, AuditLog
from stock.models import VignetteCategory
from entites.models import Entite, TypeEntite
from decimal import Decimal
from .forms import VETForm, VETVignetteFormSet, VETDocumentFormSet

from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm
from reportlab.lib import colors
from io import BytesIO

# ✅ Fonction utilitaire pour calculer le montant total (en Python)
def calculate_total_amount_for_queryset(qs):
    """
    Calcule le montant total à recouvrer pour un queryset de VET.

    Alternative à l'agrégation SQL qui ne peut pas accéder à @property.
    Charge les VET avec prefetch_related puis calcule en Python.

    Args:
        qs: QuerySet de VET objects

    Returns:
        Decimal: Montant total à recouvrer
    """
    qs = qs.prefetch_related('vignettes_assignees', 'vignettes_assignees__categorie')
    total = sum(vet.montant_total_a_recouvrer for vet in qs)
    return Decimal(str(total)) if total else Decimal('0.00')

class ExportVETPDFView(LoginRequiredMixin, View):
    """Génération d'un reçu/ordre de recette au format PDF"""
    def get(self, request, pk, *args, **kwargs):
        try:
            vet = VET.objects.prefetch_related('vignettes_assignees', 'vignettes_assignees__categorie').get(pk=pk)
            
            buffer = BytesIO()
            p = canvas.Canvas(buffer, pagesize=A4)
            width, height = A4

            # Configuration des polices (Standard)
            font_bold = "Helvetica-Bold"
            font_reg = "Helvetica"

            # En-tête
            p.setFont(font_bold, 16)
            p.drawCentredString(width/2, height - 2*cm, "RÉPUBLIQUE DU GABON")
            p.setFont(font_reg, 10)
            p.drawCentredString(width/2, height - 2.5*cm, "Union - Travail - Justice")
            p.line(width/2 - 2*cm, height - 2.7*cm, width/2 + 2*cm, height - 2.7*cm)

            # Titre du document
            p.setFont(font_bold, 14)
            p.drawCentredString(width/2, height - 4.5*cm, "ORDRE DE RECETTE / REÇU VET")
            p.setFont(font_reg, 11)
            p.drawCentredString(width/2, height - 5*cm, f"N° {vet.numero_ordre_de_recette}")

            # Infos Établissement
            p.setFont(font_bold, 12)
            p.drawString(2*cm, height - 7*cm, "INFORMATIONS DE L'ÉTABLISSEMENT")
            p.setFont(font_reg, 11)
            p.drawString(2*cm, height - 7.7*cm, f"Raison Sociale : {vet.identification_de_l_exploitant_ou_raison_sociale}")
            p.drawString(2*cm, height - 8.4*cm, f"Localisation : {vet.quartier}, {vet.zone}, {vet.region}")
            p.drawString(2*cm, height - 9.1*cm, f"Téléphone : {vet.numero_de_telephone or 'N/A'}")

            # Détails Financiers
            p.setFont(font_bold, 12)
            p.drawString(2*cm, height - 10.5*cm, "DÉTAILS DES FRAIS")
            
            y = height - 11.2*cm
            p.setFont(font_reg, 11)
            p.drawString(2*cm, y, "Libellé")
            p.drawRightString(width - 2*cm, y, "Montant (FCFA)")
            p.line(2*cm, y - 0.2*cm, width - 2*cm, y - 0.2*cm)
            
            y -= 0.8*cm
            p.drawString(2.5*cm, y, "Redevance Annuelle")
            p.drawRightString(width - 2*cm, y, f" {float(vet.montant_de_la_redevance_annuelle):,.0f}")
            
            y -= 0.6*cm
            if not vet.frais_de_dossier_payes:
                p.drawString(2.5*cm, y, "Frais de dossier")
                p.drawRightString(width - 2*cm, y, f" {float(FRAIS_DOSSIER):,.0f}")
                y -= 0.6*cm
            
            # Vignettes
            for vv in vet.vignettes_assignees.all():
                p.drawString(2.5*cm, y, f"Vignette Niveau {vv.categorie.niveau} ({float(vv.categorie.prix):,.0f} F) x{vv.quantite}")
                p.drawRightString(width - 2*cm, y, f" {float(vv.categorie.prix * vv.quantite):,.0f}")
                y -= 0.6*cm

            p.line(width - 5*cm, y - 0.2*cm, width - 2*cm, y - 0.2*cm)
            y -= 0.8*cm
            p.setFont(font_bold, 12)
            p.drawString(2*cm, y, "TOTAL À RECOUVRER")
            p.drawRightString(width - 2*cm, y, f" {float(vet.montant_total_a_recouvrer):,.0f}")

            # Statuts de paiement
            y -= 1.5*cm
            p.setFont(font_bold, 11)
            p.drawString(2*cm, y, "STATUT DES PAIEMENTS :")
            p.setFont(font_reg, 11)
            y -= 0.6*cm
            p.drawString(2.5*cm, y, f"Frais de dossier : {'PAYÉS' if vet.frais_de_dossier_payes else 'NON PAYÉS'}")
            y -= 0.6*cm
            p.drawString(2.5*cm, y, f"Redevance Annuelle : {'PAYÉE' if vet.redevance_payee else 'NON PAYÉE'}")

            # Pied de page
            p.setFont("Helvetica-Oblique", 9)
            p.drawCentredString(width/2, 2*cm, f"Document généré le {timezone.now().strftime('%d/%m/%Y à %H:%M')} - ARECOM")

            p.showPage()
            p.save()

            buffer.seek(0)
            filename = f"recu_{vet.numero_ordre_de_recette}.pdf"
            return FileResponse(buffer, as_attachment=True, filename=filename, content_type='application/pdf')
        except VET.DoesNotExist:
            return HttpResponse("Établissement non trouvé", status=404)
        except Exception as e:
            return HttpResponse(f"Erreur lors de la génération du PDF : {str(e)}", status=500)

from datetime import timedelta
import csv


class AdminRequiredMixin(UserPassesTestMixin):
    def test_func(self):
        return self.request.user.is_superuser

    def handle_no_permission(self):
        return HttpResponseForbidden("Vous n'avez pas les droits d'administration nécessaires.")


class ExportVETPDFListView(LoginRequiredMixin, AdminRequiredMixin, View):
    """Export de la liste filtrée des VET au format PDF"""
    def get(self, request, *args, **kwargs):
        # Récupération des filtres (même logique que CSV)
        now = timezone.now()
        period = request.GET.get('period', 'all')
        region = request.GET.get('region')
        qs = VET.objects.all().prefetch_related('vignettes_assignees', 'vignettes_assignees__categorie')

        if period == 'month':
            start_date = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
            qs = qs.filter(date_creation__gte=start_date)
        elif period == 'year':
            start_date = now.replace(month=1, day=1, hour=0, minute=0, second=0, microsecond=0)
            qs = qs.filter(date_creation__gte=start_date)
        
        if region:
            qs = qs.filter(region=region)

        buffer = BytesIO()
        p = canvas.Canvas(buffer, pagesize=A4)
        width, height = A4

        # En-tête
        p.setFont("Helvetica-Bold", 14)
        p.drawCentredString(width/2, height - 2*cm, "RAPPORT GLOBAL DES REDEVANCES VET")
        p.setFont("Helvetica", 10)
        p.drawCentredString(width/2, height - 2.6*cm, f"Filtres : Période={period} | Région={region or 'Toutes'}")
        p.line(1*cm, height - 3*cm, width - 1*cm, height - 3*cm)

        # Tableau
        y = height - 4*cm
        p.setFont("Helvetica-Bold", 9)
        p.drawString(1*cm, y, "N° ORDRE")
        p.drawString(4*cm, y, "RAISON SOCIALE")
        p.drawString(10*cm, y, "REGION")
        p.drawString(13*cm, y, "STATUT")
        p.drawRightString(width - 1*cm, y, "TOTAL DU (F)")
        p.line(1*cm, y - 0.2*cm, width - 1*cm, y - 0.2*cm)

        y -= 0.6*cm
        p.setFont("Helvetica", 8)
        
        for item in qs:
            if y < 2*cm: # Nouvelle page
                p.showPage()
                y = height - 2*cm
                p.setFont("Helvetica-Bold", 9)
                p.drawString(1*cm, y, "N° ORDRE")
                p.drawString(4*cm, y, "RAISON SOCIALE")
                p.drawString(10*cm, y, "REGION")
                p.drawString(13*cm, y, "STATUT")
                p.drawRightString(width - 1*cm, y, "TOTAL DU (F)")
                p.line(1*cm, y - 0.2*cm, width - 1*cm, y - 0.2*cm)
                y -= 0.6*cm
                p.setFont("Helvetica", 8)

            p.drawString(1*cm, y, str(item.numero_ordre_de_recette)[:15])
            p.drawString(4*cm, y, str(item.identification_de_l_exploitant_ou_raison_sociale)[:35])
            p.drawString(10*cm, y, str(item.region)[:15])
            p.drawString(13*cm, y, item.get_statut_display())
            p.drawRightString(width - 1*cm, y, f"{item.montant_total_a_recouvrer:,.0f}")
            y -= 0.5*cm

        p.showPage()
        p.save()

        buffer.seek(0)
        filename = f"rapport_vet_{now.strftime('%Y%m%d')}.pdf"
        return HttpResponse(buffer, content_type='application/pdf', 
                            headers={'Content-Disposition': f'attachment; filename="{filename}"'})


import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

class ExportVETExcelListView(LoginRequiredMixin, AdminRequiredMixin, View):
    """Export des données VET au format Excel (.xlsx)"""
    def get(self, request, *args, **kwargs):
        now = timezone.now()
        period = request.GET.get('period', 'all')
        region = request.GET.get('region')
        qs = VET.objects.all().prefetch_related('vignettes_assignees', 'vignettes_assignees__categorie')

        if period == 'month':
            start_date = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
            qs = qs.filter(date_creation__gte=start_date)
        elif period == 'year':
            start_date = now.replace(month=1, day=1, hour=0, minute=0, second=0, microsecond=0)
            qs = qs.filter(date_creation__gte=start_date)
        
        if region:
            qs = qs.filter(region=region)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Registre VET"

        # En-têtes avec style
        headers = [
            'N° ORDRE', 'RAISON SOCIALE', 'REGION', 'ZONE', 'QUARTIER', 
            'STATUT', 'FRAIS DOSSIER', 'REDEVANCE', 'TOTAL DU (F)'
        ]
        ws.append(headers)
        
        header_fill = PatternFill(start_color='0D6EFD', end_color='0D6EFD', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        # Données
        for item in qs:
            ws.append([
                item.numero_ordre_de_recette,
                item.identification_de_l_exploitant_ou_raison_sociale,
                item.region,
                item.zone,
                item.quartier,
                item.get_statut_display(),
                'Payé' if item.frais_de_dossier_payes else 'Non payé',
                'Payée' if item.redevance_payee else 'Non payée',
                item.montant_total_a_recouvrer  # ✅ FIX: Garder Decimal, pas float()
            ])

        # Ajuster la largeur des colonnes
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column].width = max_length + 2

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        filename = f"registre_vet_{now.strftime('%Y%m%d')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response

class ExportVETExcelDetailView(LoginRequiredMixin, View):
    """Export d'un seul VET au format Excel"""
    def get(self, request, pk, *args, **kwargs):
        vet = VET.objects.prefetch_related('vignettes_assignees', 'vignettes_assignees__categorie').get(pk=pk)
        
        wb = openpyxl.Workbook()
        ws = wb.active
        
        # Nettoyage du titre de la feuille (pas de / \ ? * : [ ]) et max 31 caractères
        safe_title = str(vet.numero_ordre_de_recette).replace('/', '-').replace('\\', '-').replace(':', '-')[:31]
        ws.title = f"Fiche {safe_title}"

        # Titre
        ws.merge_cells('A1:B1')
        ws['A1'] = f"FICHE ÉTABLISSEMENT : {vet.identification_de_l_exploitant_ou_raison_sociale}"
        ws['A1'].font = Font(size=14, bold=True)
        
        data = [
            ['N° Ordre de Recette', vet.numero_ordre_de_recette],
            ['Région', vet.region],
            ['Zone / Quartier', f"{vet.zone} / {vet.quartier}"],
            ['Statut', vet.get_statut_display()],
            ['', ''],
            ['DÉTAILS FINANCIERS', ''],
            ['Redevance Annuelle', vet.montant_de_la_redevance_annuelle],  # ✅ FIX: Garder Decimal
            ['Frais de dossier', FRAIS_DOSSIER if not vet.frais_de_dossier_payes else Decimal('0.00')],  # ✅ FIX: Garder Decimal
        ]
        
        for row in data:
            ws.append(row)
            
        ws.append(['Vignettes Assignées', ''])
        for vv in vet.vignettes_assignees.all():
            ws.append([f"Niveau {vv.categorie.niveau}", vv.categorie.prix * vv.quantite])  # ✅ FIX: Garder Decimal

        ws.append(['', ''])
        ws.append(['TOTAL À RECOUVRER', vet.montant_total_a_recouvrer])  # ✅ FIX: Garder Decimal
        
        # Style pour le total
        last_row = ws.max_row
        ws[f'A{last_row}'].font = Font(bold=True, color='0D6EFD')
        ws[f'B{last_row}'].font = Font(bold=True)

        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        # Nettoyage du nom de fichier
        safe_filename = str(vet.numero_ordre_de_recette).replace('/', '-').replace('\\', '-')
        filename = f"fiche_vet_{safe_filename}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response


class ExportVETCSVView(LoginRequiredMixin, AdminRequiredMixin, View):
    """Export des données VET au format CSV (format universel)"""
    def get(self, request, *args, **kwargs):
        # Récupération des filtres (même logique que HomeView)
        now = timezone.now()
        period = request.GET.get('period', 'all')
        region = request.GET.get('region')
        qs = VET.objects.all()

        if period == 'month':
            start_date = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
            qs = qs.filter(date_creation__gte=start_date)
        elif period == 'year':
            start_date = now.replace(month=1, day=1, hour=0, minute=0, second=0, microsecond=0)
            qs = qs.filter(date_creation__gte=start_date)
        
        if region:
            qs = qs.filter(region=region)

        # Création de la réponse CSV
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        filename = f"rapport_vet_{now.strftime('%Y%m%d_%H%M')}.csv"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        writer = csv.writer(response, delimiter=';')
        # En-têtes
        writer.writerow([
            'N° ORDRE', 'RAISON SOCIALE', 'REGION', 'ZONE', 'QUARTIER', 
            'STATUT', 'FRAIS DOSSIER PAYES', 'REDEVANCE PAYEE', 'MONTANT ANNUEL', 'TOTAL DU'
        ])

        # Données
        for item in qs:
            writer.writerow([
                item.numero_ordre_de_recette,
                item.identification_de_l_exploitant_ou_raison_sociale,
                item.region,
                item.zone,
                item.quartier,
                item.get_statut_display(),
                'OUI' if item.frais_de_dossier_payes else 'NON',
                'OUI' if item.redevance_payee else 'NON',
                item.montant_de_la_redevance_annuelle,
                item.montant_total_a_recouvrer
            ])

        return response


class VETListView(LoginRequiredMixin, ListView):
    model = VET
    template_name = "vet/vet_list.html"
    context_object_name = "vets"
    paginate_by = 20
    
    def get_queryset(self):
        qs = super().get_queryset()
        
        # Recherche textuelle
        q = self.request.GET.get("q")
        if q:
            qs = qs.filter(
                Q(identification_de_l_exploitant_ou_raison_sociale__icontains=q)
                | Q(numero_ordre_de_recette__icontains=q)
                | Q(region__icontains=q)
                | Q(zone__icontains=q)
                | Q(quartier__icontains=q)
            )
            
        # Filtre par statut
        statut = self.request.GET.get("statut")
        if statut:
            qs = qs.filter(statut=statut)
            
        # Filtre par région
        region = self.request.GET.get("region")
        if region:
            qs = qs.filter(region=region)
            
        # Filtre par frais de dossier
        frais = self.request.GET.get("frais")
        if frais == "paye":
            qs = qs.filter(frais_de_dossier_payes=True)
        elif frais == "non_paye":
            qs = qs.filter(frais_de_dossier_payes=False)
            
        # Filtre par redevance
        redevance = self.request.GET.get("redevance")
        if redevance == "paye":
            qs = qs.filter(redevance_payee=True)
        elif redevance == "non_paye":
            qs = qs.filter(redevance_payee=False)
            
        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        # Liste des régions pour le filtre (sans duplication)
        ctx["regions"] = VET.objects.order_by('region').values_list('region', flat=True).distinct()
        return ctx


class VETDetailView(LoginRequiredMixin, DetailView):
    model = VET
    template_name = "vet/vet_detail.html"
    context_object_name = "vet"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['audit_logs'] = AuditLog.objects.filter(vet=self.object).order_by('-timestamp')[:10]
        context['today'] = timezone.now().date()
        return context


class VETCreateView(LoginRequiredMixin, CreateView):
    model = VET
    form_class = VETForm
    template_name = "vet/vet_form.html"
    success_url = reverse_lazy("vet:vet_list")

    def get_context_data(self, **kwargs):
        data = super().get_context_data(**kwargs)
        if self.request.POST:
            data['vignettes'] = VETVignetteFormSet(self.request.POST)
            data['documents'] = VETDocumentFormSet(self.request.POST, self.request.FILES)
        else:
            data['vignettes'] = VETVignetteFormSet()
            data['documents'] = VETDocumentFormSet()
        return data

    def form_valid(self, form):
        context = self.get_context_data()
        vignettes = context['vignettes']
        documents = context['documents']
        if vignettes.is_valid() and documents.is_valid():
            self.object = form.save()
            vignettes.instance = self.object
            vignettes.save()
            documents.instance = self.object
            documents.save()
            # Re-sauvegarder pour mettre à jour le montant total incluant les nouvelles vignettes
            self.object.save()
            # Log audit
            AuditLog.objects.create(
                user=self.request.user,
                vet=self.object,
                action='create',
                details=f"Création de l'établissement {self.object.numero_ordre_de_recette}"
            )
            return super().form_valid(form)
        else:
            return self.render_to_response(self.get_context_data(form=form))


class VETUpdateView(LoginRequiredMixin, UpdateView):
    model = VET
    form_class = VETForm
    template_name = "vet/vet_form.html"
    success_url = reverse_lazy("vet:vet_list")

    def get_context_data(self, **kwargs):
        data = super().get_context_data(**kwargs)
        if self.request.POST:
            data['vignettes'] = VETVignetteFormSet(self.request.POST, instance=self.object)
            data['documents'] = VETDocumentFormSet(self.request.POST, self.request.FILES, instance=self.object)
        else:
            data['vignettes'] = VETVignetteFormSet(instance=self.object)
            data['documents'] = VETDocumentFormSet(instance=self.object)
        return data

    def form_valid(self, form):
        context = self.get_context_data()
        vignettes = context['vignettes']
        documents = context['documents']
        if vignettes.is_valid() and documents.is_valid():
            self.object = form.save()
            vignettes.instance = self.object
            vignettes.save()
            documents.instance = self.object
            documents.save()
            # Re-sauvegarder pour mettre à jour le montant total incluant les nouvelles vignettes
            self.object.save()
            # Log audit (simple diff simulation)
            AuditLog.objects.create(
                user=self.request.user,
                vet=self.object,
                action='update',
                details=f"Modification de l'établissement {self.object.numero_ordre_de_recette}"
            )
            return super().form_valid(form)
        else:
            return self.render_to_response(self.get_context_data(form=form))


class VETDeleteView(LoginRequiredMixin, AdminRequiredMixin, DeleteView):
    model = VET
    template_name = "vet/vet_confirm_delete.html"
    success_url = reverse_lazy("vet:vet_list")


class HomeView(LoginRequiredMixin, TemplateView):
    template_name = "home.html"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        now = timezone.now()
        
        # Filtres temporels
        period = self.request.GET.get('period', 'all')
        qs = VET.objects.all()
        
        if period == 'month':
            start_date = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
            qs = qs.filter(date_creation__gte=start_date)
        elif period == 'year':
            start_date = now.replace(month=1, day=1, hour=0, minute=0, second=0, microsecond=0)
            qs = qs.filter(date_creation__gte=start_date)
            
        # Filtres géographiques
        region = self.request.GET.get('region')
        if region:
            qs = qs.filter(region=region)
            
        # KPI de base
        ctx["total_vet"] = qs.count()
        ctx["actifs"] = qs.filter(statut="actif").count()
        ctx["inactifs"] = qs.filter(statut="inactif").count()
        
        # KPI Financiers
        # Montant total des redevances annuelles (dues ou payées)
        total_redevance = qs.aggregate(total=Sum('montant_de_la_redevance_annuelle'))['total'] or 0

        # Montant total à recouvrer (calcul via @property avec prefetch_related)
        # Note: montant_total_a_recouvrer inclut redevance + frais dossier (si non payé) + vignettes
        total_a_recouvrer = calculate_total_amount_for_queryset(qs)

        # Montants collectés (Précis : redevance payée + frais dossier payés)
        collectes_frais = qs.filter(frais_de_dossier_payes=True).count() * FRAIS_DOSSIER
        collectes_redevance = qs.filter(redevance_payee=True).aggregate(total=Sum('montant_de_la_redevance_annuelle'))['total'] or 0
        collectes = collectes_frais + collectes_redevance

        # Retards de paiement (basé sur la date d'expiration)
        en_retard = qs.filter(date_d_expiration_de_la_redevance_annuelle__lt=now.date(), statut='actif').count()

        ctx["total_redevance"] = total_redevance
        ctx["total_a_recouvrer"] = total_a_recouvrer
        ctx["taux_recouvrement"] = round((collectes / total_a_recouvrer * 100), 1) if total_a_recouvrer > 0 else 0
        ctx["en_retard"] = en_retard

        # Comparaisons par région (Top 5)
        # ✅ OPTIMISATION #1: Charger UNE FOIS en Python au lieu de N requêtes
        # Grouper les VET par région
        regions_data = {}
        for vet in qs.prefetch_related('vignettes_assignees', 'vignettes_assignees__categorie'):
            region = vet.region
            if region not in regions_data:
                regions_data[region] = []
            regions_data[region].append(vet)

        # Calculer count et total pour chaque région
        regions_stats = []
        for region, vets in regions_data.items():
            region_total = sum(v.montant_total_a_recouvrer for v in vets)
            regions_stats.append({
                'region': region,
                'count': len(vets),
                'total': float(region_total)
            })

        # Trier par total descendant et prendre les top 5
        stats_region = sorted(regions_stats, key=lambda x: x['total'], reverse=True)[:5]
        ctx["stats_region"] = stats_region
        
        # Alertes de stock bas
        ctx['low_stock_alerts'] = VignetteCategory.objects.filter(stock_actuel__lte=F('seuil_alerte'))
        
        # Liste des régions pour le filtre (sans duplication et triée)
        ctx["regions"] = VET.objects.order_by('region').values_list('region', flat=True).distinct()
        ctx["current_period"] = period
        ctx["current_region"] = region
        
        # --- Données pour les Graphiques ---
        # 1. Évolution mensuelle (6 derniers mois)
        six_months_ago = now - timedelta(days=180)

        # Charger les VET avec données liées pour calcul du montant
        vets_by_month = {}
        for vet in VET.objects.filter(date_creation__gte=six_months_ago) \
                .prefetch_related('vignettes_assignees', 'vignettes_assignees__categorie'):
            month_key = vet.date_creation.strftime('%Y-%m')
            if month_key not in vets_by_month:
                vets_by_month[month_key] = []
            vets_by_month[month_key].append(vet)

        # Calculer les totaux par mois
        monthly_stats = []
        for month_str in sorted(vets_by_month.keys()):
            vets = vets_by_month[month_str]
            total_month = sum(v.montant_total_a_recouvrer for v in vets)
            monthly_stats.append({
                'month': month_str,
                'total': total_month
            })
        
        ctx["monthly_labels"] = [stat['month'].replace('-', ' / ') for stat in monthly_stats]
        ctx["monthly_data"] = [float(stat['total']) for stat in monthly_stats]

        # 2. Répartition par région (pour le camembert)
        ctx["region_labels"] = [stat['region'] for stat in stats_region]
        ctx["region_data"] = [float(stat['total']) for stat in stats_region]
        
        # 3. Répartition par Niveau de Vignette
        vignette_stats = VETVignette.objects.filter(vet__in=qs).values('categorie__niveau') \
            .annotate(total_qty=Sum('quantite')) \
            .order_by('categorie__niveau')
        
        ctx["vignette_labels"] = [f"Niveau {stat['categorie__niveau']}" for stat in vignette_stats]
        ctx["vignette_data"] = [stat['total_qty'] for stat in vignette_stats]

        # 4. Statut des Paiements
        ctx["payment_stats"] = {
            "redevance_payee": qs.filter(redevance_payee=True).count(),
            "redevance_due": qs.filter(redevance_payee=False).count(),
            "frais_payes": qs.filter(frais_de_dossier_payes=True).count(),
            "frais_dus": qs.filter(frais_de_dossier_payes=False).count(),
        }

        # --- Stats des ENTITÉS (Distributeurs, Réseaux, Cybercafés) ---
        ctx["total_entites"] = Entite.objects.count()

        # Grouper par type d'entité
        entites_par_type = Entite.objects.values('type_entite__nom').annotate(
            count=Count('id'),
            total_redevance=Sum('frais_exploitation_annuel')
        ).order_by('-count')
        ctx["entites_par_type"] = entites_par_type

        # Dernières entités pour affichage
        ctx["latest_entites"] = Entite.objects.select_related('type_entite').order_by('-created_at')[:5]

        # Total des redevances entités
        ctx["total_redevance_entites"] = Entite.objects.aggregate(
            total=Sum('frais_exploitation_annuel')
        )['total'] or Decimal('0.00')

        return ctx


class VETMapView(LoginRequiredMixin, AdminRequiredMixin, TemplateView):
    """Carte interactive avec données GPS (admin seulement)"""
    template_name = "vet/vet_maps.html"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        
        # On récupère les VET avec leurs vignettes assignées
        vets = (
            VET.objects.exclude(latitude__isnull=True)
            .exclude(longitude__isnull=True)
            .prefetch_related('vignettes_assignees', 'vignettes_assignees__categorie')
        )
        
        vet_points = []
        for v in vets:
            # On construit une chaîne résumant les vignettes pour l'affichage
            vignettes_summary = ", ".join([
                f"N{vv.categorie.niveau}" for vv in v.vignettes_assignees.all()
            ])
            
            vet_points.append({
                "numero": v.numero,
                "identification_de_l_exploitant_ou_raison_sociale": v.identification_de_l_exploitant_ou_raison_sociale,
                "latitude": v.latitude,
                "longitude": v.longitude,
                "region": v.region,
                "zone": v.zone,
                "quartier": v.quartier,
                "statut": v.statut,
                "frais_de_dossier_payes": v.frais_de_dossier_payes,
                "montant_de_la_redevance_annuelle": float(v.montant_de_la_redevance_annuelle),
                "montant_total_a_recouvrer": float(v.montant_total_a_recouvrer),
                "vignettes_summary": vignettes_summary or "Aucune",
                "dossier_suivi_par": v.dossier_suivi_par,
                "date_d_expiration_de_la_redevance_annuelle": v.date_d_expiration_de_la_redevance_annuelle.isoformat() if v.date_d_expiration_de_la_redevance_annuelle else "",
                "numero_ordre_de_recette": v.numero_ordre_de_recette,
            })
            
        ctx["vet_points"] = vet_points
        ctx["regions"] = VET.objects.order_by('region').values_list('region', flat=True).distinct()
        return ctx
